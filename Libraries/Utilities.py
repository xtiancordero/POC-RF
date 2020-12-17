from openpyxl import load_workbook
from robot.api import logger
from pathlib import Path
from pynput.keyboard import Key, Controller
import pyperclip
import datetime


class Utilities:
	def get_column_text_index(self, ws, colHeadingText):
		"""
            get column index of given column header text
        """

		counter = 0
		col_num = 0
		for col in ws.iter_cols(ws.min_column, ws.max_column, min_row=1, max_row=1):
			for cell in col:
				counter += 1
				if cell.value == colHeadingText:
					col_num = counter
					break
		return col_num

	def get_row_text_index(self, ws, rowHeadingText):
		"""
            get row index of given 1st row text
        """
		counter = 0
		row_num = 0
		for row in ws.iter_rows(ws.min_row, ws.max_row, min_col=2, max_col=2):
			for cell in row:
				counter += 1
				if cell.value == rowHeadingText:
					row_num = counter
					break
		return row_num

	def get_cell_text(self, path, sheet, row_heading, column_heading, length=None):
		"""
            get cell text using row heading and column heading index and trim as needed
            Example:
           | Get Cell Text  | excelPATH |  sheet1  | Admin  |  UserName |
           | Get Cell Text  | excelPATH |  sheet1  | Admin  |  UserName |   3
        """
		try:
			wb = load_workbook(path)
			ws = wb[sheet]

			column = self.get_column_text_index(ws, column_heading)
			row = self.get_row_text_index(ws, row_heading)

			cell_value = ws.cell(row, column).value

			if cell_value is None or '':
				str_cell_value = ''
			else:
				if length is None:
					str_cell_value = str(cell_value).strip()
				else:
					cell_value = str(cell_value).strip()
					str_cell_value = cell_value[:int(length)]

			return str_cell_value

		except Exception as e:
			logger.warn(str(e) + ':{}{} Row Header:' + str(row_heading) + ' or Column Header:'
						+ str(column_heading) + ' not found')

	def get_credential(self, path, sheet, row_heading, column_heading, length=None):
		"""
            get cell text using row heading and column heading index and trim as needed
            Example:
           | Get Cell Text  | excelPATH |  sheet1  | Admin  |  UserName |
           | Get Cell Text  | excelPATH |  sheet1  | Admin  |  UserName |   3
        """
		try:
			wb = load_workbook(path)
			ws = wb[sheet]
			column = self.get_column_text_index(ws, column_heading)
			row = self.get_row_text_index(ws, row_heading)

			cell_value = ws.cell(row, column).value

			if cell_value is None or '':
				str_cell_value = ''
			else:
				if length is None:
					str_cell_value = str(cell_value).strip()
				else:
					cell_value = str(cell_value).strip()
					str_cell_value = cell_value[:int(length)]

			return str_cell_value

		except Exception as e:
			logger.warn(str(e) + ':{}{} Row Header:' + str(row_heading) + ' or Column Header:'
						+ str(column_heading) + ' not found')

	def write_cell_text(self, path, sheet, row_heading, column_heading, value):
		"""
            write value to excel

            Example:
            |Write Cell Text | excelPATH |  sheet1  | Admin  |  UserName | FOO |
        :param path:
        :param sheet:
        :param row_heading:
        :param column_heading:
        :return:
        """

		wb = load_workbook(path)
		# wb = load_workbook(path)
		ws = wb[sheet]

		column = self.get_column_text_index(ws, column_heading)
		row = self.get_row_text_index(ws, row_heading)

		ws.cell(row, column).value = value
		wb.save(path)

	def format_ptf(self, ptfNumber):
		"""
            changes the format of portfolio to application accepted format
                 e.g. ptfNumber = 556677081
                 Change Ptf Number Format  |  ptfNo    |   accepted format

                 *output will be  '556677-08-1'
        """
		ptfList = [int(x) for x in (ptfNumber)]
		ptfList.insert(6, '-')
		ptfList.insert(9, '-')
		newFormat = ''.join(str(e) for e in ptfList)
		return newFormat

	def change_amount_number_format(self, amount):

		"""
            changes the format of amount to be retrieved in excel
            amount should not be a string
            e.g.
                2000 -> 2,000
                123456 -> 123,456
        :param amount:
        :return:
        """
		new_format = "{:,}".format(float(amount))
		return new_format

	def change_date_format(self, input_date):
		"""
            changes the date format
            e.g.
            e.g.
                11/29/2018  ->  112918
        :param input_date:
        :return:
        """
		string_list = input_date.split('\n')
		date_list = string_list[1].split('/')
		mon = date_list[0]
		date = date_list[1]
		year = date_list[2]
		return mon + date + year[2:4]

	def get_audit_detail_value(self, list, detailname, delimeter):
		"""
            For finding the values in the Audit Report Details Page
        :param list:
        :param detailname:
        :param delimeter:
        :return:
        """
		for value in list:
			if detailname in value:
				detaillist = value.split(delimeter)
				return detaillist[1]

	def remove_white_space(self, message):
		"""
             removes white spaces on a string
                  e.g. mystring = 'Here is  the   text   I      want   '
                    Remove White Space | myString
                *output will be  'Here is the text i want'
            """
		newString = " ".join(message.split())
		return newString

	def remove_special_character(self, message):
		"""

        :param message:
        :return:
        """

		newString = message.replace("/", "").replace("\\", "").replace(":", "").replace("-", "").replace("<",
																										 "").replace(
			">", "").replace(",", "")
		return newString

	def compare_strings(self, string1, string2, ignore_case=False):
		result = False
		if ignore_case:
			string1 = string1.lower()
			string2 = string2.lower()

		if string1 == string2:
			result = True
		return result

	def get_excel_row_count(self, path, sheet):
		"""
        :param path: excel file path
        :param sheet: sheet name
        :return: maximum number of row of excel sheet
        """

		wb = load_workbook(path)
		# wb = load_workbook(path)
		ws = wb[sheet]
		maxRow = ws.max_row
		return maxRow

	def get_excel_column_count(self, path, sheet):
		"""
        :param path: excel file path
        :param sheet: sheet name
        :return: maximum number of column of excel sheet
        """

		wb = load_workbook(path)
		# wb = load_workbook(path)
		ws = wb[sheet]
		maxColumn = ws.max_column
		return maxColumn

	def remove_all_spaces(self, string):
		newString = string.replace(" ", "")
		return newString

	def remove_space_accountHolder(self, string):
		newString = string[:6] + string[7:]
		return newString

	def add_space_accountHolder(self, string):
		newString = string[:6] + " " + string[6:]
		return newString

	def clean_list(self, list):
		newString = list.replace("[", "").replace("]", "").replace("u", "").replace(",", "\n").replace("\'", "")
		return newString

	def get_date_year(self, date):
		"""    get year from given date
			e.g.
			e.g.
				2020-10-01   ->  2020"""
		string_list = list(date.split(" "))
		year = string_list[2]
		return year

	def get_date_month(self, date):
		"""    get year from given date
			e.g.
			e.g.
				2020-10-01   ->  2020"""
		string_list = list(date.split(" "))
		month = string_list[0]
		return month

	def get_date_day(self, date):
		"""    get year from given date
			e.g.
			e.g.
				2020-10-01   ->  2020"""
		string_list = list(date.split(" "))
		day = string_list[1].replace(",", "")
		return day

	def convert_country_code(self, country):
		ph = 'Philippines'
		if country == 'PH':
			country = ph
		return country

	def use_date_format_one(self, date):
		'''format given date to Month D, Yr ie Jan 1, 2020'''
		month = self.get_date_month(date)
		day = self.get_date_day(date)
		year = self.get_date_year(date)
		
		newmonth = month[0:3]
		newdate = newmonth+" "+day+","+" "+year
		return newdate

	def use_date_format_two(self, date):
		'''format given date to Month D, Yr ie 04-Dec-2020'''
		month = self.get_date_month(date)
		day = self.get_date_day(date)
		year = self.get_date_year(date)
		
		newdate = day+"-"+month+"-"+year
		return newdate




	def press_enter_key (self):
		keyboard = Controller()
		keyboard.press(Key.enter)
		keyboard.release(Key.enter)

	def press_ctrl_f (self):
		keyboard = Controller()
		keyboard.press(Key.ctrl)
		keyboard.press('f')
		keyboard.release(Key.ctrl)
		keyboard.release('f')

	def press_ctrl_v (self):
		keyboard = Controller()
		keyboard.press(Key.ctrl)
		keyboard.press('v')
		keyboard.release(Key.ctrl)
		keyboard.release('v')

	def press_f3 (self):
		keyboard = Controller()
		keyboard.press(Key.f3)
		keyboard.release(Key.f3)
		
	def copy_data (self, data):
		pyperclip.copy(data)
		pyperclip.paste()
		return data
	
	def paste_data (self):
		pyperclip.paste()
		return None


